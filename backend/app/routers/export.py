"""
API routes for exporting data to PDF and Excel
"""
from fastapi import APIRouter, Depends, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional, List
import io
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from ..database import get_db
from ..models import US, InventarioMateriali, Pottery, Site

router = APIRouter(prefix="/export", tags=["Export"])


def create_excel_workbook(data: list, columns: list, title: str) -> io.BytesIO:
    """Create an Excel workbook from data"""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_idx, column in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column['label'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, column in enumerate(columns, 1):
            value = getattr(row_data, column['field'], None)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

    # Auto-adjust column widths
    for col_idx, column in enumerate(columns, 1):
        max_length = len(str(column['label']))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_pdf_document(data: list, columns: list, title: str) -> io.BytesIO:
    """Create a PDF document from data"""
    output = io.BytesIO()

    # Use landscape for more columns
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1.5*cm,
        bottomMargin=1*cm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=20
    )

    elements = []

    # Title
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Prepare table data
    table_data = [[col['label'] for col in columns]]

    for row_data in data:
        row = []
        for col in columns:
            value = getattr(row_data, col['field'], None)
            if value is None:
                value = ""
            elif len(str(value)) > 50:
                value = str(value)[:47] + "..."
            row.append(str(value))
        table_data.append(row)

    # Calculate column widths
    available_width = landscape(A4)[0] - 2*cm
    col_width = available_width / len(columns)
    col_widths = [col_width] * len(columns)

    # Create table
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Header style
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),

        # Data style
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # Borders
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E9EDF4')]),
    ]))

    elements.append(table)

    # Build PDF
    doc.build(elements)
    output.seek(0)
    return output


# US Export columns
US_COLUMNS = [
    {'field': 'sito', 'label': 'Site'},
    {'field': 'area', 'label': 'Area'},
    {'field': 'us', 'label': 'SU'},
    {'field': 'd_stratigrafica', 'label': 'Strat. Definition'},
    {'field': 'd_interpretativa', 'label': 'Interpretation'},
    {'field': 'periodo_iniziale', 'label': 'Period'},
    {'field': 'fase_iniziale', 'label': 'Phase'},
    {'field': 'datazione', 'label': 'Dating'},
    {'field': 'descrizione', 'label': 'Description'},
    {'field': 'interpretazione', 'label': 'Interpretation Notes'},
]

# Materials Export columns
MATERIALI_COLUMNS = [
    {'field': 'sito', 'label': 'Site'},
    {'field': 'numero_inventario', 'label': 'Inv. No.'},
    {'field': 'tipo_reperto', 'label': 'Type'},
    {'field': 'definizione', 'label': 'Definition'},
    {'field': 'area', 'label': 'Area'},
    {'field': 'us', 'label': 'SU'},
    {'field': 'nr_cassa', 'label': 'Box No.'},
    {'field': 'luogo_conservazione', 'label': 'Storage Location'},
    {'field': 'stato_conservazione', 'label': 'Condition'},
    {'field': 'datazione_reperto', 'label': 'Dating'},
    {'field': 'totale_frammenti', 'label': 'Tot. Fragments'},
    {'field': 'peso', 'label': 'Weight (g)'},
]

# Pottery Export columns - updated for pottery_table structure
POTTERY_COLUMNS = [
    {'field': 'sito', 'label': 'Site'},
    {'field': 'id_number', 'label': 'ID'},
    {'field': 'area', 'label': 'Area'},
    {'field': 'us', 'label': 'SU'},
    {'field': 'form', 'label': 'Form'},
    {'field': 'specific_form', 'label': 'Specific Form'},
    {'field': 'material', 'label': 'Material'},
    {'field': 'fabric', 'label': 'Fabric'},
    {'field': 'ware', 'label': 'Ware'},
    {'field': 'box', 'label': 'Box'},
    {'field': 'qty', 'label': 'Quantity'},
    {'field': 'note', 'label': 'Notes'},
]


@router.get("/us/excel")
async def export_us_excel(
    sito: Optional[str] = None,
    area: Optional[str] = None,
    periodo: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Export stratigraphic units to Excel"""
    query = db.query(US)

    if sito:
        query = query.filter(US.sito == sito)
    if area:
        query = query.filter(US.area == area)
    if periodo:
        query = query.filter(US.periodo_iniziale == periodo)

    data = query.order_by(US.sito, US.area, US.us).all()

    if not data:
        raise HTTPException(status_code=404, detail="No data to export")

    title = f"SU_{sito or 'all'}_{datetime.now().strftime('%Y%m%d')}"
    output = create_excel_workbook(data, US_COLUMNS, title)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={title}.xlsx"}
    )


@router.get("/us/pdf")
async def export_us_pdf(
    sito: Optional[str] = None,
    area: Optional[str] = None,
    periodo: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Export stratigraphic units to PDF"""
    query = db.query(US)

    if sito:
        query = query.filter(US.sito == sito)
    if area:
        query = query.filter(US.area == area)
    if periodo:
        query = query.filter(US.periodo_iniziale == periodo)

    data = query.order_by(US.sito, US.area, US.us).all()

    if not data:
        raise HTTPException(status_code=404, detail="No data to export")

    title = f"Stratigraphic Units - {sito or 'All Sites'}"
    output = create_pdf_document(data, US_COLUMNS[:8], title)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=SU_{sito or 'all'}_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )


@router.get("/materiali/excel")
async def export_materiali_excel(
    sito: Optional[str] = None,
    nr_cassa: Optional[int] = None,
    luogo_conservazione: Optional[str] = None,
    tipo_reperto: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Export materials inventory to Excel"""
    query = db.query(InventarioMateriali)

    if sito:
        query = query.filter(InventarioMateriali.sito == sito)
    if nr_cassa:
        query = query.filter(InventarioMateriali.nr_cassa == nr_cassa)
    if luogo_conservazione:
        query = query.filter(InventarioMateriali.luogo_conservazione == luogo_conservazione)
    if tipo_reperto:
        query = query.filter(InventarioMateriali.tipo_reperto == tipo_reperto)

    data = query.order_by(InventarioMateriali.sito, InventarioMateriali.numero_inventario).all()

    if not data:
        raise HTTPException(status_code=404, detail="No data to export")

    title = f"Materials_{sito or 'all'}_{datetime.now().strftime('%Y%m%d')}"
    output = create_excel_workbook(data, MATERIALI_COLUMNS, title)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={title}.xlsx"}
    )


@router.get("/materiali/pdf")
async def export_materiali_pdf(
    sito: Optional[str] = None,
    nr_cassa: Optional[int] = None,
    luogo_conservazione: Optional[str] = None,
    tipo_reperto: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Export materials inventory to PDF"""
    query = db.query(InventarioMateriali)

    if sito:
        query = query.filter(InventarioMateriali.sito == sito)
    if nr_cassa:
        query = query.filter(InventarioMateriali.nr_cassa == nr_cassa)
    if luogo_conservazione:
        query = query.filter(InventarioMateriali.luogo_conservazione == luogo_conservazione)
    if tipo_reperto:
        query = query.filter(InventarioMateriali.tipo_reperto == tipo_reperto)

    data = query.order_by(InventarioMateriali.sito, InventarioMateriali.numero_inventario).all()

    if not data:
        raise HTTPException(status_code=404, detail="No data to export")

    title = f"Materials Inventory - {sito or 'All Sites'}"
    output = create_pdf_document(data, MATERIALI_COLUMNS[:10], title)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Materials_{sito or 'all'}_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )


@router.get("/materiali/search/excel")
async def export_materials_search_excel(
    ids: str = Query(..., description="Comma-separated list of material IDs"),
    db: Session = Depends(get_db)
):
    """Export specific materials by IDs to Excel (for search results)"""
    id_list = [int(id.strip()) for id in ids.split(',') if id.strip().isdigit()]

    if not id_list:
        raise HTTPException(status_code=400, detail="No valid IDs provided")

    data = db.query(InventarioMateriali).filter(
        InventarioMateriali.id_invmat.in_(id_list)
    ).order_by(InventarioMateriali.numero_inventario).all()

    if not data:
        raise HTTPException(status_code=404, detail="No data to export")

    title = f"Materials_Search_{datetime.now().strftime('%Y%m%d_%H%M')}"
    output = create_excel_workbook(data, MATERIALI_COLUMNS, title)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={title}.xlsx"}
    )


@router.get("/materiali/search/pdf")
async def export_materials_search_pdf(
    ids: str = Query(..., description="Comma-separated list of material IDs"),
    db: Session = Depends(get_db)
):
    """Export specific materials by IDs to PDF (for search results)"""
    id_list = [int(id.strip()) for id in ids.split(',') if id.strip().isdigit()]

    if not id_list:
        raise HTTPException(status_code=400, detail="No valid IDs provided")

    data = db.query(InventarioMateriali).filter(
        InventarioMateriali.id_invmat.in_(id_list)
    ).order_by(InventarioMateriali.numero_inventario).all()

    if not data:
        raise HTTPException(status_code=404, detail="No data to export")

    title = f"Materials Search Results - {len(data)} items"
    output = create_pdf_document(data, MATERIALI_COLUMNS[:10], title)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Materials_Search_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"}
    )


@router.get("/pottery/excel")
async def export_pottery_excel(
    sito: Optional[str] = None,
    form: Optional[str] = None,
    material: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Export pottery to Excel"""
    query = db.query(Pottery)

    if sito:
        query = query.filter(Pottery.sito == sito)
    if form:
        query = query.filter(Pottery.form == form)
    if material:
        query = query.filter(Pottery.material == material)

    data = query.order_by(Pottery.sito, Pottery.id_number).all()

    if not data:
        raise HTTPException(status_code=404, detail="No data to export")

    title = f"Pottery_{sito or 'all'}_{datetime.now().strftime('%Y%m%d')}"
    output = create_excel_workbook(data, POTTERY_COLUMNS, title)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={title}.xlsx"}
    )


@router.get("/pottery/pdf")
async def export_pottery_pdf(
    sito: Optional[str] = None,
    form: Optional[str] = None,
    material: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Export pottery to PDF"""
    query = db.query(Pottery)

    if sito:
        query = query.filter(Pottery.sito == sito)
    if form:
        query = query.filter(Pottery.form == form)
    if material:
        query = query.filter(Pottery.material == material)

    data = query.order_by(Pottery.sito, Pottery.id_number).all()

    if not data:
        raise HTTPException(status_code=404, detail="No data to export")

    title = f"Pottery - {sito or 'All Sites'}"
    output = create_pdf_document(data, POTTERY_COLUMNS[:9], title)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Pottery_{sito or 'all'}_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )


@router.get("/inventory/summary/excel")
async def export_inventory_summary_excel(
    sito: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Export detailed inventory summary (boxes and storage) to Excel"""
    query = db.query(InventarioMateriali)
    if sito:
        query = query.filter(InventarioMateriali.sito == sito)

    all_materials = query.all()

    if not all_materials:
        raise HTTPException(status_code=404, detail="No data to export")

    # Group by storage and box
    storage_data = defaultdict(lambda: defaultdict(list))
    for mat in all_materials:
        storage = mat.luogo_conservazione or "Not specified"
        box = mat.nr_cassa or 0
        storage_data[storage][box].append(mat)

    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    storage_font = Font(bold=True, size=12)
    storage_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row = 1

    # Title
    ws_summary.cell(row=row, column=1, value=f"Warehouse Inventory Summary - {sito or 'All Sites'}")
    ws_summary.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 1
    ws_summary.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    row += 2

    # Overall stats
    total_items = len(all_materials)
    total_boxes = len(set((m.luogo_conservazione, m.nr_cassa) for m in all_materials if m.nr_cassa))
    total_weight = sum(m.peso or 0 for m in all_materials)
    total_fragments = sum(m.totale_frammenti or 0 for m in all_materials)

    ws_summary.cell(row=row, column=1, value="Overall Statistics")
    ws_summary.cell(row=row, column=1).font = storage_font
    row += 1
    ws_summary.cell(row=row, column=1, value=f"Total Items: {total_items}")
    row += 1
    ws_summary.cell(row=row, column=1, value=f"Total Boxes: {total_boxes}")
    row += 1
    ws_summary.cell(row=row, column=1, value=f"Total Weight: {total_weight/1000:.2f} kg")
    row += 1
    ws_summary.cell(row=row, column=1, value=f"Total Fragments: {total_fragments}")
    row += 2

    # Category breakdown
    by_type = defaultdict(int)
    for mat in all_materials:
        tipo = mat.tipo_reperto or "Unknown"
        by_type[tipo] += 1

    ws_summary.cell(row=row, column=1, value="Items by Category")
    ws_summary.cell(row=row, column=1).font = storage_font
    row += 1

    for col, header in enumerate(["Category", "Count"], 1):
        cell = ws_summary.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1

    for tipo, count in sorted(by_type.items(), key=lambda x: -x[1]):
        ws_summary.cell(row=row, column=1, value=tipo).border = thin_border
        ws_summary.cell(row=row, column=2, value=count).border = thin_border
        row += 1

    row += 2

    # Storage location breakdown
    for storage_name, boxes in sorted(storage_data.items()):
        # Storage location header
        ws_summary.cell(row=row, column=1, value=f"Storage: {storage_name}")
        ws_summary.cell(row=row, column=1).font = storage_font
        ws_summary.cell(row=row, column=1).fill = storage_fill
        ws_summary.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

        # Box headers
        headers = ["Box No.", "Items", "Categories", "Weight (g)", "Fragments", "SU Range"]
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        # Box data
        for box_name, materials in sorted(boxes.items()):
            types = set(m.tipo_reperto for m in materials if m.tipo_reperto)
            total_weight = sum(m.peso or 0 for m in materials)
            total_fragments = sum(m.totale_frammenti or 0 for m in materials)
            us_list = sorted(set(str(m.us) for m in materials if m.us))
            us_range = f"{us_list[0]}-{us_list[-1]}" if len(us_list) > 1 else (us_list[0] if us_list else "-")

            ws_summary.cell(row=row, column=1, value=box_name).border = thin_border
            ws_summary.cell(row=row, column=2, value=len(materials)).border = thin_border
            ws_summary.cell(row=row, column=3, value=", ".join(sorted(types))).border = thin_border
            ws_summary.cell(row=row, column=4, value=round(total_weight, 2)).border = thin_border
            ws_summary.cell(row=row, column=5, value=total_fragments).border = thin_border
            ws_summary.cell(row=row, column=6, value=us_range).border = thin_border
            row += 1

        row += 1  # Empty row between storage locations

    # Auto-adjust column widths
    for col in range(1, 7):
        ws_summary.column_dimensions[get_column_letter(col)].width = 18

    # Create detailed sheet with all items
    ws_detail = wb.create_sheet("All Items")

    # Headers
    for col_idx, column in enumerate(MATERIALI_COLUMNS, 1):
        cell = ws_detail.cell(row=1, column=col_idx, value=column['label'])
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Data sorted by storage, box, inventory number
    sorted_materials = sorted(all_materials, key=lambda m: (
        m.luogo_conservazione or "",
        m.nr_cassa or 0,
        m.numero_inventario or 0
    ))

    for row_idx, mat in enumerate(sorted_materials, 2):
        for col_idx, column in enumerate(MATERIALI_COLUMNS, 1):
            value = getattr(mat, column['field'], None)
            cell = ws_detail.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Auto-adjust column widths
    for col_idx, column in enumerate(MATERIALI_COLUMNS, 1):
        ws_detail.column_dimensions[get_column_letter(col_idx)].width = 15

    ws_detail.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    title = f"Inventory_Summary_{sito or 'all'}_{datetime.now().strftime('%Y%m%d')}"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={title}.xlsx"}
    )


@router.get("/inventory/summary/pdf")
async def export_inventory_summary_pdf(
    sito: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Export inventory summary to PDF"""
    query = db.query(InventarioMateriali)
    if sito:
        query = query.filter(InventarioMateriali.sito == sito)

    all_materials = query.all()

    if not all_materials:
        raise HTTPException(status_code=404, detail="No data to export")

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1.5*cm,
        bottomMargin=1*cm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    section_style = ParagraphStyle(
        'Section',
        parent=styles['Heading2'],
        fontSize=12,
        spaceBefore=15,
        spaceAfter=10
    )

    elements = []

    # Title
    elements.append(Paragraph(f"Warehouse Inventory Summary - {sito or 'All Sites'}", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Overall statistics
    total_items = len(all_materials)
    total_boxes = len(set((m.luogo_conservazione, m.nr_cassa) for m in all_materials if m.nr_cassa))
    total_weight = sum(m.peso or 0 for m in all_materials)

    stats_data = [
        ["Total Items", str(total_items)],
        ["Total Boxes", str(total_boxes)],
        ["Total Weight", f"{total_weight/1000:.2f} kg"],
        ["Storage Locations", str(len(set(m.luogo_conservazione for m in all_materials if m.luogo_conservazione)))]
    ]

    stats_table = Table(stats_data, colWidths=[150, 100])
    stats_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 20))

    # Category breakdown
    by_type = defaultdict(int)
    for mat in all_materials:
        tipo = mat.tipo_reperto or "Unknown"
        by_type[tipo] += 1

    elements.append(Paragraph("Items by Category", section_style))

    cat_data = [["Category", "Count"]]
    for tipo, count in sorted(by_type.items(), key=lambda x: -x[1]):
        cat_data.append([tipo, str(count)])

    cat_table = Table(cat_data, colWidths=[200, 80])
    cat_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E9EDF4')]),
    ]))
    elements.append(cat_table)
    elements.append(Spacer(1, 20))

    # Storage breakdown
    storage_data = defaultdict(lambda: defaultdict(list))
    for mat in all_materials:
        storage = mat.luogo_conservazione or "Not specified"
        box = mat.nr_cassa or 0
        storage_data[storage][box].append(mat)

    elements.append(Paragraph("Boxes by Storage Location", section_style))

    for storage_name, boxes in sorted(storage_data.items()):
        elements.append(Paragraph(f"<b>{storage_name}</b>", styles['Normal']))

        box_data = [["Box", "Items", "Categories", "Weight"]]
        for box_name, materials in sorted(boxes.items()):
            types = set(m.tipo_reperto for m in materials if m.tipo_reperto)
            total_wt = sum(m.peso or 0 for m in materials)
            box_data.append([
                str(box_name),
                str(len(materials)),
                ", ".join(sorted(types))[:40],
                f"{total_wt:.0f}g"
            ])

        box_table = Table(box_data, colWidths=[60, 50, 300, 60])
        box_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E9EDF4')]),
        ]))
        elements.append(box_table)
        elements.append(Spacer(1, 10))

    doc.build(elements)
    output.seek(0)

    title = f"Inventory_Summary_{sito or 'all'}_{datetime.now().strftime('%Y%m%d')}"

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={title}.pdf"}
    )
